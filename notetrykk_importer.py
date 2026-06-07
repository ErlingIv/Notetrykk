"""
Notetrykk — Supabase Importer
Reads notetrykk_import.xlsx and inserts into Supabase.

Usage:
    pip install requests openpyxl
    python notetrykk_importer.py

The Excel file must be in the same folder as this script.
Columns (row 1 = header, data from row 2):
    A: Person       — "Last, First" or "A | B" for multiple
    B: Rolle        — Komponist / Arrangør / Utgiver  (| separated for multiple)
    C: Tittel       — title, "/subtitle", "(Lyricist)" extraction automatic
    D: Besetning    — legacy_code, or "code1 | code2" for multiple (first known used)
    E: Forlag       — publisher name
    F: Platenr      — plate number
    G: Utgivelsesår — year, "1880 - 1889", "1905 ca" etc.
    H: Måned        — month string
    I: Periodikum   — "Nordisk Musik-Tidende 1890 h.10" etc.
    J: Kommentar    — notes (stored in composition_notes)
"""

import re
import sys
from pathlib import Path
import requests
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────
SUPABASE_URL = "https://lmsqnyssnxsiibnyguxy.supabase.co"
SUPABASE_KEY = "sb_secret_REDACTED_ROTATE_ME"   # paste your sb_secret_... key here

EXCEL_FILE   = Path(__file__).parent / "notetrykk_import.xlsx"

HEADERS = {
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "apikey":        SUPABASE_KEY,
    "Content-Type":  "application/json",
    "Prefer":        "return=representation",
}

# ── Role mapping ──────────────────────────────────────────────────────────────
ROLE_MAP = {
    "Komponist": "Composer",
    "Arrangør":  "Arranger",
    "Utgiver":   "Editor",
    "Redaktør":  "Editor",
}

# ── Caches (populated at runtime) ─────────────────────────────────────────────
INSTRUMENTATION_CACHE = {}   # legacy_code  → instrumentation_id
PUBLISHER_CACHE       = {}   # name         → publisher_id
PERSON_CACHE          = {}   # "Last, First" → person_id
PERIODICAL_CACHE      = {}   # title        → periodical_id
ISSUE_CACHE           = {}   # (pid, ref)   → issue_id
COMPLETE_PERSONS      = set()  # person_ids with is_complete=TRUE — rows skipped in pre-flight
NEW_PERSONS           = {}     # raw_name → person_id for persons inserted this session
UNKNOWN_CODES         = set()

# ── Instrumentation code aliases ──────────────────────────────────────────────
# Maps source codes that don't exactly match legacy_code in the DB
INSTRUMENTATION_ALIASES = {
    "2 vn":        "2vn",
    "s-salmodikon": "s-salm",
    "s-2 fl":      "s-2fl",
}

# ── Publication type codes ─────────────────────────────────────────────────────
# These are not instrumentation — they describe publication format.
# Mapped to publication_type field on composition.
PUBLICATION_TYPE_CODES = {
    "sangbok":      "sangbok",
    "skolesangbok": "skolesangbok",
    "koralbok":     "koralbok",
    "litur":        "liturgisk",
    "skole":        "skole",
    "skole s":      "skole",      # school material, sang-og-piano instrumentation
    "skole org":    "skole",      # school material, organ instrumentation
    "skole p":      "skole",      # school material, piano instrumentation
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def api_get(path):
    r = requests.get(f"{SUPABASE_URL}/rest/v1/{path}", headers=HEADERS)
    r.raise_for_status()
    return r.json()

def api_post(table, payload, extra_prefer=""):
    prefer = "return=representation"
    if extra_prefer:
        prefer += "," + extra_prefer
    r = requests.post(
        f"{SUPABASE_URL}/rest/v1/{table}",
        headers={**HEADERS, "Prefer": prefer},
        json=payload,
    )
    if r.status_code not in (200, 201):
        raise RuntimeError(f"POST {table} failed {r.status_code}: {r.text[:300]}")
    return r.json()[0]


def load_instrumentation():
    rows = api_get("instrumentation?select=instrumentation_id,legacy_code")
    for row in rows:
        if row["legacy_code"]:
            INSTRUMENTATION_CACHE[row["legacy_code"]] = row["instrumentation_id"]
    print(f"  Loaded {len(INSTRUMENTATION_CACHE)} instrumentation codes")


def parse_year(raw):
    raw = (raw or "").strip()
    if not raw:
        return None, None, None
    if re.match(r"^\d{4} - \d{4}$", raw):
        parts = raw.split(" - ")
        return parts[0], parts[1], "range"
    if raw.endswith(" ca"):
        return raw[:-3].strip(), None, "circa"
    if raw.endswith("c"):
        return raw[:-1], None, "circa"
    if raw.endswith("f"):
        return raw[:-1], None, "before"
    if re.match(r"^\d{3}\*$", raw):
        return raw[:-1] + "0", None, "decade"
    if re.match(r"^\d{4}$", raw):
        return raw, None, "exact"
    return raw, None, None   # fallback


def parse_title(raw):
    """Returns (title, subtitle, lyricists_list)."""
    raw = (raw or "").strip()
    subtitle = None
    if "/" in raw:
        idx = raw.index("/")
        title_part = raw[:idx].strip()
        subtitle   = raw[idx + 1:].strip()
    else:
        title_part = raw

    lyricists = []
    for text in [title_part, subtitle]:
        if not text:
            continue
        for match in re.findall(r"\(([^)]+)\)", text):
            words = match.strip().split()
            if 1 <= len(words) <= 4 and not any(c.isdigit() for c in match):
                lyricists.append(match.strip())

    return title_part, subtitle, lyricists


# Month number → Norwegian abbreviation (for Den norske Lyra YYMM parsing)
DNL_MONTH_MAP = {
    "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
    "05": "Mai", "06": "Jun", "07": "Jul", "08": "Aug",
    "09": "Sep", "10": "Okt", "11": "Nov", "12": "Des"
}

def parse_periodikum(raw):
    """Returns (periodical_title, raw_reference) or (None, None)."""
    raw = (raw or "").strip()
    if not raw:
        return None, None
    # "Title YYYY h.XYZ" — year between title and ref (NMT style)
    # e.g. "Nordisk Musik-Tidende 1890 h.10" → title="Nordisk Musik-Tidende", ref="h.10"
    m = re.match(r"^(.+?)\s+\d{4}\s+((?:NR |TrR )?h\.\w+)$", raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    # "Title YYYY ekstrabilag" — year + label, no h. ref → no parseable ref
    m = re.match(r"^(.+?)\s+\d{4}\s+\w+$", raw)
    if m:
        return m.group(1).strip(), None
    # Standard: "Title h.XYZ" or "Title NR h.02"
    m = re.match(r"^(.+?)\s+((?:NR |TrR )?h\.\w+)$", raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    # Den norske Lyra: "Title YYMM" — 4-digit number alone
    m = re.match(r"^(.+?)\s+(\d{4})$", raw)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return raw, None


def get_or_create_publisher(name):
    name = name.strip()
    if not name:
        return None
    if name in PUBLISHER_CACHE:
        return PUBLISHER_CACHE[name]
    # Look up existing publisher before inserting
    rows = api_get(f"publisher?publisher_name=eq.{requests.utils.quote(name)}&select=publisher_id")
    if rows:
        PUBLISHER_CACHE[name] = rows[0]["publisher_id"]
        return rows[0]["publisher_id"]
    row = api_post("publisher", {
        "publisher_name":    name,
        "city":              "Christiania",
        "is_self_published": name.lower() == "eget",
        "is_unknown":        name.lower() == "ukjent",
    })
    PUBLISHER_CACHE[name] = row["publisher_id"]
    return row["publisher_id"]


def get_or_create_person(raw_name):
    raw_name = raw_name.strip()
    if not raw_name:
        return None
    if raw_name in PERSON_CACHE:
        return PERSON_CACHE[raw_name]

    if raw_name.lower() == "anon":
        last, first, ptype = "Anon", None, "anon"
    elif raw_name.lower().startswith("anon ("):
        last, first, ptype = raw_name, None, "anon"
    elif raw_name.upper() == "NN":
        last, first, ptype = "NN", None, "collection"
    elif "," in raw_name:
        parts  = raw_name.split(",", 1)
        last   = parts[0].strip()
        first  = parts[1].strip() or None
        ptype  = "person"
    else:
        last, first, ptype = raw_name, None, "person"

    # Look up existing person before inserting
    params = f"last_name=eq.{requests.utils.quote(last)}"
    if first:
        params += f"&first_name=eq.{requests.utils.quote(first)}"
    else:
        params += "&first_name=is.null"
    rows = api_get(f"person?{params}&select=person_id,is_complete")
    if rows:
        pid = rows[0]["person_id"]
        PERSON_CACHE[raw_name] = pid
        if rows[0].get("is_complete"):
            COMPLETE_PERSONS.add(pid)
        return pid

    payload = {"last_name": last, "person_type": ptype}
    if first:
        payload["first_name"] = first

    row = api_post("person", payload)
    PERSON_CACHE[raw_name] = row["person_id"]
    NEW_PERSONS[raw_name] = row["person_id"]
    return row["person_id"]


def get_periodical_id(title):
    if title in PERIODICAL_CACHE:
        return PERIODICAL_CACHE[title]
    rows = api_get(f"periodical?title=eq.{requests.utils.quote(title)}&select=periodical_id")
    if not rows:
        print(f"  WARNING: periodical not found: '{title}'")
        return None
    PERIODICAL_CACHE[title] = rows[0]["periodical_id"]
    return rows[0]["periodical_id"]


def get_or_create_issue(periodical_id, raw_ref, original_string=None):
    key = (periodical_id, raw_ref)
    if key in ISSUE_CACHE:
        return ISSUE_CACHE[key]

    volume = issue_number = series_label = division = section = None

    # Den norske Lyra: 4-digit YYMM e.g. "2308" = Aug 1823
    m_dnl = re.match(r"^(\d{2})(\d{2})$", raw_ref)

    # Standard h. pattern: optional series label + h. + digits
    m_h = re.match(r"^(NR |TrR )?h\.(\d+)$", raw_ref)

    if m_dnl:
        # YYMM: year = 1800 + YY, month stored as Norwegian abbreviation
        yy   = m_dnl.group(1)
        mm   = m_dnl.group(2)
        # volume repurposed as full year, issue_number as month number
        volume       = 1800 + int(yy)
        issue_number = int(mm)
    elif m_h:
        series_label = m_h.group(1).strip() if m_h.group(1) else None
        digits = m_h.group(2)
        if len(digits) <= 2:
            # e.g. h.02 → issue_number=2
            issue_number = int(digits)
        elif len(digits) == 3:
            # e.g. h.101 → volume=1, issue_number=1
            volume       = int(digits[0])
            issue_number = int(digits[1:])
        elif len(digits) == 4:
            # MLM h.0102 → volume=1, issue_number=2
            # MLM h.1201 → division=1, section=2, issue_number=1
            first = int(digits[0])
            if first == 0:
                volume       = int(digits[1])
                issue_number = int(digits[2:])
            else:
                division     = first
                section      = int(digits[1])
                issue_number = int(digits[2:])

    # Store full original string if available, otherwise just the parsed ref
    payload = {"periodical_id": periodical_id, "raw_reference": original_string or raw_ref}
    if series_label: payload["series_label"] = series_label
    if volume:       payload["volume"]        = volume
    if issue_number: payload["issue_number"]  = issue_number
    if division:     payload["division"]      = division
    if section:      payload["section"]       = section

    row = api_post("periodical_issue", payload)
    ISSUE_CACHE[key] = row["issue_id"]
    return row["issue_id"]


# ── Duplicate detection ──────────────────────────────────────────────────────
def find_existing_composition(title, publisher_id, plate_number, year_issued, persons, instrumentation_id=None):
    """
    Returns composition_id if a matching composition already exists in the DB, else None.
    Only checks against the DATABASE — never suppresses rows within a single paste batch.
    Match strategy:
      1. title + publisher_id + year_issued + plate_number (when present)
      2. + instrumentation_id (when present) — catches same title/year in different arrangements
      3. Verify primary composer matches
    """
    title_enc = requests.utils.quote(title)
    params = f"title=eq.{title_enc}"
    if publisher_id:
        params += f"&publisher_id=eq.{publisher_id}"
    if year_issued:
        params += f"&year_issued=eq.{year_issued}"
    if plate_number:
        params += f"&plate_number=eq.{plate_number}"
    if instrumentation_id:
        params += f"&instrumentation_id=eq.{instrumentation_id}"

    candidates = api_get(f"composition?{params}&select=composition_id")
    if not candidates:
        return None

    # Get primary composer from incoming record
    primary_composer = next(
        (pname for pname, role in persons if role == "Komponist"), None
    )
    if not primary_composer:
        return candidates[0]["composition_id"]

    primary_person_id = PERSON_CACHE.get(primary_composer)
    if not primary_person_id:
        parts = primary_composer.split(",", 1)
        last  = parts[0].strip()
        rows  = api_get(
            f"person?last_name=eq.{requests.utils.quote(last)}&select=person_id"
        )
        if not rows:
            return candidates[0]["composition_id"]
        primary_person_id = rows[0]["person_id"]

    # Check each candidate for a matching composer
    for candidate in candidates:
        cid = candidate["composition_id"]
        cp_rows = api_get(
            f"composition_person?composition_id=eq.{cid}"
            f"&person_id=eq.{primary_person_id}&role=eq.Composer&select=id"
        )
        if cp_rows:
            return cid

    return None


# ── Excel reader ──────────────────────────────────────────────────────────────

def read_excel(path):
    # Use read_only=False so we can write the imported_id stamp back
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    records = []
    for row_num in range(2, ws.max_row + 1):
        def col(i):
            v = ws.cell(row=row_num, column=i+1).value
            return str(v).strip() if v is not None else ""

        # Skip if already imported (col K has a value)
        imported_id = ws.cell(row=row_num, column=11).value
        if imported_id:
            continue

        persons_raw  = col(0)
        roles_raw    = col(1)
        title_raw    = col(2)
        besetning    = col(3)
        publisher    = col(4)
        plate        = col(5)
        year         = col(6)
        month        = col(7)
        periodikum   = col(8)
        kommentar    = col(9)

        # Skip fully empty rows
        if not any([persons_raw, title_raw, publisher]):
            continue

        # Multiple persons/roles separated by " | "
        person_names = [p.strip() for p in persons_raw.split("|") if p.strip()]
        role_strings = [r.strip() for r in roles_raw.split("|")   if r.strip()]
        if not role_strings:
            role_strings = ["Komponist"]
        if not person_names:
            continue

        persons = []
        for idx, pname in enumerate(person_names):
            role_str = role_strings[idx] if idx < len(role_strings) else role_strings[-1]
            persons.append((pname, role_str))

        records.append({
            "row_num":    row_num,
            "persons":    persons,
            "title_raw":  title_raw,
            "besetning":  besetning,
            "publisher":  publisher,
            "plate":      plate,
            "year":       year,
            "month":      month or None,
            "periodikum": periodikum,
            "kommentar":  kommentar,
            "raw_source": {
                "person":     persons_raw,
                "rolle":      roles_raw,
                "tittel":     title_raw,
                "besetning":  besetning,
                "forlag":     publisher,
                "platenr":    plate,
                "aar":        year,
                "maaned":     month or "",
                "periodikum": periodikum,
                "kommentar":  kommentar,
            },
        })
    return wb, ws, records


# ── Main import ───────────────────────────────────────────────────────────────

def import_records():
    if not EXCEL_FILE.exists():
        print(f"ERROR: Excel file not found: {EXCEL_FILE}")
        sys.exit(1)

    print(f"Reading {EXCEL_FILE.name}...")
    wb, ws, records = read_excel(EXCEL_FILE)
    print(f"  {len(records)} new rows to import (stamped rows skipped)")

    print("Loading instrumentation codes from Supabase...")
    load_instrumentation()
    print()

    inserted = skipped = errors = 0
    skipped_periodicals = []   # (title_raw, periodikum_string) for end-of-run summary

    for rec in records:
        title_raw = rec["title_raw"]
        try:
            # 1. Title / subtitle / lyricists
            title, subtitle, lyricists = parse_title(title_raw)

            # 1b. Pre-resolve persons — skip row instantly if all are marked complete
            resolved_persons = []
            for pname, role_str in rec["persons"]:
                pid = get_or_create_person(pname)
                resolved_persons.append((pname, role_str, pid))
            if resolved_persons and all(pid in COMPLETE_PERSONS for _, _, pid in resolved_persons if pid):
                print(f"  SKIP (all persons complete) {title_raw[:60]}")
                skipped += 1
                continue

            # 2. Year
            year_issued, year_issued_end, year_qualifier = parse_year(rec["year"])

            # 3. Instrumentation + publication_type resolution
            besetning_raw = rec["besetning"].strip()
            instrumentation_id  = None
            raw_instrumentation = None
            publication_type    = None

            # Check if the whole besetning string is a known publication type
            if besetning_raw in PUBLICATION_TYPE_CODES:
                publication_type = PUBLICATION_TYPE_CODES[besetning_raw]
                # Some publication codes imply an instrumentation too
                implied = {"skole s": "s-p", "skole org": "org", "skole p": "p2"}
                if besetning_raw in implied:
                    instrumentation_id = INSTRUMENTATION_CACHE.get(implied[besetning_raw])
            else:
                # Try each | or space separated token as an instrumentation code
                # Apply aliases for source codes that don't match legacy_code exactly
                codes = [c.strip() for c in besetning_raw.replace("|", " ").split()]
                for code in codes:
                    resolved = INSTRUMENTATION_ALIASES.get(code, code)
                    if resolved in INSTRUMENTATION_CACHE:
                        instrumentation_id = INSTRUMENTATION_CACHE[resolved]
                        break
                # Also try the whole besetning_raw as an alias (e.g. "2 vn")
                if not instrumentation_id:
                    resolved = INSTRUMENTATION_ALIASES.get(besetning_raw, besetning_raw)
                    if resolved in INSTRUMENTATION_CACHE:
                        instrumentation_id = INSTRUMENTATION_CACHE[resolved]
                if not instrumentation_id and besetning_raw:
                    raw_instrumentation = besetning_raw
                    UNKNOWN_CODES.add(besetning_raw)

            # 4. Publisher
            publisher_id = get_or_create_publisher(rec["publisher"]) if rec["publisher"] else None

            # 5. Composition
            comp = {"title": title}
            if subtitle:             comp["subtitle"]            = subtitle
            if publisher_id:         comp["publisher_id"]        = publisher_id
            if instrumentation_id:   comp["instrumentation_id"]  = instrumentation_id
            # Always store original besetning string regardless of whether it was mapped
            comp["raw_instrumentation"] = rec["besetning"] if rec["besetning"] else None
            if rec["plate"]:         comp["plate_number"]        = rec["plate"]
            if year_issued:          comp["year_issued"]         = year_issued
            if year_issued_end:      comp["year_issued_end"]     = year_issued_end
            if year_qualifier:       comp["year_qualifier"]      = year_qualifier
            if rec["month"]:         comp["month"]               = rec["month"]
            if rec["kommentar"]:     comp["composition_notes"]   = rec["kommentar"]
            if publication_type:     comp["publication_type"]    = publication_type
            comp["raw_source"] = rec["raw_source"]

            # Check for existing composition before inserting
            existing_id = find_existing_composition(
                title, publisher_id, rec["plate"], year_issued, rec["persons"],
                instrumentation_id=instrumentation_id
            )
            if existing_id:
                print(f"  SKIP [{existing_id:4d}] (exists in DB) {title[:60]}")
                ws.cell(row=rec["row_num"], column=11).value = existing_id
                skipped += 1
                continue

            comp_row = api_post("composition", comp)
            composition_id = comp_row["composition_id"]

            # Stamp the Excel row with the composition_id
            ws.cell(row=rec["row_num"], column=11).value = composition_id

            # 6. Persons
            for idx, (pname, role_str, person_id) in enumerate(resolved_persons):
                if not person_id:
                    continue
                api_post("composition_person", {
                    "composition_id": composition_id,
                    "person_id":      person_id,
                    "role":           ROLE_MAP.get(role_str, "Composer"),
                    "is_primary":     (idx == 0),
                })

            # 7. Lyricists extracted from title
            for lyr in lyricists:
                person_id = get_or_create_person(lyr)
                if person_id:
                    api_post("composition_person", {
                        "composition_id": composition_id,
                        "person_id":      person_id,
                        "role":           "Lyricist",
                        "is_primary":     False,
                    })

            # 8. Periodikum
            if rec["periodikum"]:
                per_title, raw_ref = parse_periodikum(rec["periodikum"])
                if not raw_ref:
                    skipped_periodicals.append((title_raw, rec["periodikum"], "no parseable ref"))
                elif per_title and raw_ref:
                    per_id = get_periodical_id(per_title)
                    if not per_id:
                        skipped_periodicals.append((title_raw, rec["periodikum"], "periodical not in DB"))
                    if per_id:
                        issue_id = get_or_create_issue(per_id, raw_ref, original_string=rec["periodikum"])
                        if issue_id:
                            api_post("composition_issue", {
                                "composition_id": composition_id,
                                "issue_id":       issue_id,
                            }, extra_prefer="resolution=merge-duplicates")

            inserted += 1
            print(f"  OK  [{composition_id:4d}] {title[:65]}")

        except Exception as e:
            print(f"  ERR  '{title_raw[:60]}': {e}")
            errors += 1

    wb.save(EXCEL_FILE)
    print(f"\n─── Done: {inserted} inserted, {skipped} skipped, {errors} errors ───")
    print(f"    Excel file updated with imported_id stamps")
    if UNKNOWN_CODES:
        print("\nUnknown instrumentation codes → stored as raw_instrumentation:")
        for c in sorted(UNKNOWN_CODES):
            print(f"  {c!r}")

    if skipped_periodicals:
        print(f"\nSkipped periodical links ({len(skipped_periodicals)}) — manual follow-up needed:")
        for title_r, periodi, reason in skipped_periodicals:
            print(f"  [{reason}] {periodi!r}  ←  {title_r[:50]}")

    # ── Post-import: offer to mark composers as complete ─────────────────────
    if NEW_PERSONS:
        print("\n─── Newly imported persons this session ───")
        person_list = []
        for raw_name, person_id in sorted(NEW_PERSONS.items(), key=lambda x: x[0]):
            # Count compositions in DB for this person
            try:
                rows = api_get(
                    f"composition_person?person_id=eq.{person_id}&role=eq.Composer&select=composition_id"
                )
                count = len(rows)
            except:
                count = '?'
            print(f"  [{person_id:4d}] {raw_name} ({count} komposisjoner i DB)")
            person_list.append((raw_name, person_id))

        print("\nMark any composers as is_complete = TRUE?")
        print("Enter person_ids separated by commas, or press Enter to skip:")
        try:
            answer = input("> ").strip()
            if answer:
                ids = [x.strip() for x in answer.split(',') if x.strip().isdigit()]
                if ids:
                    for pid in ids:
                        name = next((n for n, i in person_list if str(i) == pid), pid)
                        try:
                            requests.patch(
                                f"{SUPABASE_URL}/rest/v1/person?person_id=eq.{pid}",
                                headers=HEADERS,
                                json={"is_complete": True}
                            )
                            print(f"  ✓ Marked {name} (id={pid}) as complete")
                        except Exception as e:
                            print(f"  ERROR marking {name}: {e}")
        except (EOFError, KeyboardInterrupt):
            pass


if __name__ == "__main__":
    import_records()
